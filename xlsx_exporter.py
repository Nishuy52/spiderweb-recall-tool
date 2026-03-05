from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from models import Person

INITIATOR_FILL  = PatternFill("solid", start_color="FFD700")   # gold
UNAVAIL_FILL    = PatternFill("solid", start_color="D3D3D3")   # grey
HEADER_FILL     = PatternFill("solid", start_color="1F3864")   # dark navy
HEADER_FONT     = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT       = Font(name="Arial", size=10)
BOLD_BODY       = Font(name="Arial", size=10, bold=True)
CENTER          = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT            = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin = Side(style="thin", color="AAAAAA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

PLATOON_COLORS = [
    "E8F4FD", "FEF9E7", "E9F7EF", "F9EBEA", "EBE9F7",
    "E8F8F5", "FDF2F8", "F0F3F4", "FFF3E0", "E3F2FD",
]

def platoon_fill(platoon: str, platoon_map: dict) -> PatternFill:
    idx = platoon_map.get(platoon, 0) % len(PLATOON_COLORS)
    return PatternFill("solid", start_color=PLATOON_COLORS[idx])

def export_xlsx(people: list[Person], output_path: str):
    wb = Workbook()

    # --- Sheet 1: Call Assignments ---
    ws = wb.active
    ws.title = "Call Assignments"
    ws.freeze_panes = "A2"

    headers = ["Rank", "Name", "Platoon", "People to Call", "Called By"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    # Build platoon colour map
    platoons = sorted(set(p.platoon for p in people))
    platoon_map = {plt: i for i, plt in enumerate(platoons)}

    all_people = sorted(people, key=lambda p: (p.platoon, p.rank_level, p.name))

    for p in all_people:
        calls_str    = ", ".join(f"{c.rank} {c.name}" for c in p.calls)
        calledby_str = ", ".join(f"{c.rank} {c.name}" for c in p.called_by)

        row = [p.rank, p.name, p.platoon, calls_str, calledby_str]
        ws.append(row)
        r = ws.max_row

        if not p.available:
            fill = UNAVAIL_FILL
        elif p.is_initiator:
            fill = INITIATOR_FILL
        else:
            fill = platoon_fill(p.platoon, platoon_map)

        for col, _ in enumerate(row, 1):
            cell = ws.cell(row=r, column=col)
            cell.fill = fill
            cell.font = BOLD_BODY if p.is_initiator else BODY_FONT
            cell.alignment = LEFT if col >= 4 else CENTER
            cell.border = BORDER

    col_widths = [8, 18, 10, 42, 42]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 2: Summary Stats ---
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    available = [p for p in people if p.available]
    reached   = [p for p in available if p.called_by or p.is_initiator]

    stats = [
        ("Total Personnel",        len(people)),
        ("Available",              len(available)),
        ("Unavailable",            len(people) - len(available)),
        ("Initiators",             len([p for p in people if p.is_initiator])),
        ("Personnel Reachable",    len(reached)),
        ("Personnel Unreachable",  len(available) - len(reached)),
        ("Coverage %",             f"{len(reached)/len(available)*100:.1f}%" if available else "N/A"),
    ]

    ws2.append(["Metric", "Value"])
    for col in range(1, 3):
        cell = ws2.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for label, val in stats:
        ws2.append([label, val])
        r = ws2.max_row
        for col in range(1, 3):
            cell = ws2.cell(row=r, column=col)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    # Legend
    ws2.append([])
    ws2.append(["Legend", ""])
    legend_row = ws2.max_row
    ws2.cell(row=legend_row, column=1).font = BOLD_BODY

    legend = [
        ("Initiator (OC/CSM/2IC)", "FFD700"),
        ("Unavailable",            "D3D3D3"),
        ("Active (by platoon colour)", PLATOON_COLORS[0]),
    ]
    for label, color in legend:
        ws2.append([label, ""])
        r = ws2.max_row
        ws2.cell(row=r, column=1).fill = PatternFill("solid", start_color=color)
        ws2.cell(row=r, column=1).font = BODY_FONT
        ws2.cell(row=r, column=1).border = BORDER

    wb.save(output_path)